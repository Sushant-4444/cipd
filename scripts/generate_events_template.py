"""
Generates an xlsx template that clients can fill, one event per row.
Output: events-template.xlsx in the project root.

Run:  python3 scripts/generate_events_template.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

OUT_PATH = Path(__file__).resolve().parent.parent / "events-template.xlsx"

# ── Column schema ────────────────────────────────────────────────────────────
# Each tuple: (header, width, help_comment, example_value)
COLUMNS = [
    ("ID",                    8,  "Unique integer per event. Use 1, 2, 3, …", 1),
    ("Type",                  18, "Pick from dropdown: Webinar, Workshop, Conference, Community Meetup, Summit.", "Webinar"),
    ("Title",                 38, "Full event title shown on cards and the detail page.", "AI Tools for HR Teams"),
    ("Date",                  20, "Human-readable date, e.g. 15 April 2026.", "15 April 2026"),
    ("Time",                  28, "Time range with timezone, e.g. 6:30 PM – 7:30 PM IST.", "6:30 PM – 7:30 PM IST"),
    ("Location",              22, "City, venue, or 'Online'. e.g. 'Mumbai' or 'Hybrid · IIIT Delhi + Online'.", "Online"),
    ("Format",                28, "Detailed format for detail page, e.g. 'Live Webinar (Zoom)', 'In-Person Workshop', 'Hybrid (IIIT Delhi + Online)'.", "Live Webinar (Zoom)"),
    ("Status",                14, "Pick from dropdown: Upcoming, Featured, Past. Only one event should be 'Featured'.", "Upcoming"),
    ("Featured",              12, "Pick from dropdown: Yes / No. Only ONE event should be Yes — it appears at the top.", "No"),
    ("Summary",               60, "Short 1-2 line description shown on the event card.", "Learn practical AI workflows for recruitment, onboarding, and performance tracking."),
    ("Full Description",      80, "Longer paragraph (3-5 sentences) shown on the detail page 'About This Event' section.",
        "This webinar brings together HR practitioners and AI researchers to explore practical workflows for integrating AI tools into daily HR operations. From automated resume screening to intelligent performance dashboards, you'll see real demos and leave with actionable templates."),
    ("Topics (| separated)",  50, "Tags shown as pills on the detail page. Separate with a pipe: AI in HR | LLMs | HR Copilot",
        "AI-Powered Recruitment | LLM-Based Performance Tracking | HR Copilot Demo"),
    ("Capacity",              12, "Total seats available. UPCOMING events only. Leave blank for past events.", 500),
    ("Registered",            14, "Registrations so far. UPCOMING events only. Drives the capacity progress bar.", 342),
    ("Prerequisites",         50, "Eligibility / requirements shown in a yellow callout. e.g. 'None — open to all.' or '3+ years of HR experience recommended.'", "None — open to all HR professionals."),
    ("Agenda (one per line)", 80, "One agenda item per line. Format on each line:\n  TIME | TITLE | SPEAKER\nExample:\n  6:30 PM | Welcome & Opening | CiPD Team\n  6:40 PM | AI in Recruitment | Dr. Ananya Rao",
        "6:30 PM | Welcome & Opening | CiPD Team\n6:40 PM | AI-Powered Recruitment | Dr. Ananya Rao\n7:00 PM | Performance Tracking with LLMs | Vikram Sehgal\n7:30 PM | Q&A & Wrap-up | Panel"),
    ("Speakers (one per line)", 60, "One speaker per line. Format on each line:\n  NAME | ROLE\nExample:\n  Dr. Ananya Rao | Head of People Analytics, Zeta Corp",
        "Dr. Ananya Rao | Head of People Analytics, Zeta Corp\nVikram Sehgal | Senior ML Engineer, TalentAI"),
    ("Past Action",           20, "PAST events only. Pick from dropdown: View Highlights, Watch Recording, Read Recap. Sets the link label on past event cards.", ""),
    ("Highlights (one per line)", 70, "PAST events only. Bullet points shown in 'Key Takeaways'. One per line.", ""),
    ("Image Filenames (| separated)", 60, "Photo filenames the team will upload to /public/events/. Pipe-separated.\nExample: webinar-ai-1.jpg | webinar-ai-2.jpg",
        "webinar-ai-1.jpg | webinar-ai-2.jpg | webinar-ai-3.jpg"),
    ("Notes for Web Team",    40, "Optional. Anything else the web team should know (special graphics, sponsors, livestream URL, etc.).", ""),
]

# Second example: a past event
PAST_EXAMPLE = [
    2,
    "Conference",
    "Annual HR Strategy Forum 2025",
    "14 December 2025",
    "10:00 AM – 4:00 PM IST",
    "Delhi",
    "In-Person Conference",
    "Past",
    "No",
    "Strategic discussions on HR operating models, talent pipelines, and people analytics adoption.",
    "The 2025 Annual HR Strategy Forum brought together 200+ senior HR leaders for a day of strategic discussions on operating models, talent pipelines, and analytics adoption. Attendees walked away with actionable frameworks and new industry connections.",
    "HR Operating Models | Talent Pipelines | People Analytics Adoption",
    "",
    "",
    "",
    "10:00 AM | Opening & Industry Overview | Moderator\n10:45 AM | HR Operating Models for Scale | Panel\n12:00 PM | Talent Pipeline Strategy | Speaker\n2:00 PM | Analytics Adoption Roadmap | Workshop",
    "",
    "View Highlights",
    "200+ senior HR leaders attended\n3 keynote speakers from Fortune 500 companies\n78% plan to adopt people analytics in 2026\nCommunity-voted Best Session: Talent Pipeline Strategy",
    "forum-2025-1.jpg | forum-2025-2.jpg | forum-2025-3.jpg",
    "",
]

# ── Build workbook ───────────────────────────────────────────────────────────
wb = Workbook()

# ===== Sheet 1: Events =====
ws = wb.active
ws.title = "Events"

header_fill = PatternFill(start_color="00BFA5", end_color="00BFA5", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
example_fill = PatternFill(start_color="F2FBF9", end_color="F2FBF9", fill_type="solid")
example_font = Font(name="Calibri", size=10, italic=True, color="555555")
data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

# Header row
for col_idx, (header, width, help_text, _example) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border
    if help_text:
        cell.comment = Comment(help_text, "CiPD Web Team")
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 36
ws.freeze_panes = "A2"

# Example rows (italic, lightly shaded — show clients exactly what to fill)
example_upcoming = [c[3] for c in COLUMNS]
for row_idx, example_data in enumerate([example_upcoming, PAST_EXAMPLE], start=2):
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = example_fill
        cell.font = example_font
        cell.alignment = data_align
        cell.border = thin_border

# Mark example rows with a label in row height
ws.row_dimensions[2].height = 110
ws.row_dimensions[3].height = 110

# Reserve empty rows for client to fill
for row_idx in range(4, 24):
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.alignment = data_align
        cell.border = thin_border

# ── Data validations (dropdowns) ─────────────────────────────────────────────
def col_letter(name):
    for i, (h, *_rest) in enumerate(COLUMNS):
        if h == name:
            return get_column_letter(i + 1)
    return None

dv_type = DataValidation(
    type="list",
    formula1='"Webinar,Workshop,Conference,Community Meetup,Summit"',
    allow_blank=True,
    showDropDown=False,
)
dv_type.error = "Pick one of: Webinar, Workshop, Conference, Community Meetup, Summit."
dv_type.errorTitle = "Invalid Type"
ws.add_data_validation(dv_type)
dv_type.add(f"{col_letter('Type')}2:{col_letter('Type')}200")

dv_status = DataValidation(
    type="list",
    formula1='"Upcoming,Featured,Past"',
    allow_blank=True,
)
dv_status.error = "Pick one of: Upcoming, Featured, Past."
dv_status.errorTitle = "Invalid Status"
ws.add_data_validation(dv_status)
dv_status.add(f"{col_letter('Status')}2:{col_letter('Status')}200")

dv_featured = DataValidation(
    type="list",
    formula1='"Yes,No"',
    allow_blank=True,
)
dv_featured.error = "Pick Yes or No. Only one event should be Yes."
dv_featured.errorTitle = "Invalid Featured Value"
ws.add_data_validation(dv_featured)
dv_featured.add(f"{col_letter('Featured')}2:{col_letter('Featured')}200")

dv_past_action = DataValidation(
    type="list",
    formula1='"View Highlights,Watch Recording,Read Recap"',
    allow_blank=True,
)
ws.add_data_validation(dv_past_action)
dv_past_action.add(f"{col_letter('Past Action')}2:{col_letter('Past Action')}200")

# ===== Sheet 2: Instructions =====
ws2 = wb.create_sheet("Instructions")
ws2.column_dimensions["A"].width = 110

instructions = [
    ("CiPD Events — Submission Template",          "title"),
    ("",                                            "blank"),
    ("How to use this file",                       "h2"),
    ("• Add ONE event per row in the 'Events' sheet.",                "p"),
    ("• Rows 2 and 3 are pre-filled examples — keep them for reference, or delete before sending back.",  "p"),
    ("• Hover over each column header (red triangle) to see field-specific help.",  "p"),
    ("• For multi-line fields (Agenda, Speakers, Highlights), put one item per line within the cell. Press Alt+Enter (Windows) or Option+Enter (Mac) for a new line inside a cell.",  "p"),
    ("• For pipe-separated fields (Topics, Image Filenames), use the | character between values. e.g. 'AI in HR | LLMs | Workforce'.",  "p"),
    ("",                                            "blank"),
    ("Field reference",                            "h2"),
    ("ID — unique integer (1, 2, 3, …).",          "p"),
    ("Type — must be one of the dropdown values: Webinar, Workshop, Conference, Community Meetup, Summit.",  "p"),
    ("Title — appears on cards and the detail page hero.",  "p"),
    ("Date — human-readable, e.g. '15 April 2026'.",  "p"),
    ("Time — include timezone, e.g. '6:30 PM – 7:30 PM IST'.",  "p"),
    ("Location — short label for cards (city, 'Online', or 'Hybrid · venue').",  "p"),
    ("Format — longer label for detail page (e.g. 'Live Webinar (Zoom)', 'In-Person Workshop', 'Hybrid · IIIT Delhi + Online').",  "p"),
    ("Status — Upcoming / Featured / Past. Only one event may be Featured at a time.",  "p"),
    ("Featured — Yes for the single hero event; No for everything else.",  "p"),
    ("Summary — 1-2 lines, shown on cards.",        "p"),
    ("Full Description — longer paragraph for the detail page.",  "p"),
    ("Topics — pipe-separated tag list, shown as pills.",  "p"),
    ("Capacity / Registered — used to draw the capacity bar on upcoming events. Leave blank for past events.",  "p"),
    ("Prerequisites — eligibility line shown in a yellow callout. Leave blank if none.",  "p"),
    ("Agenda — one line per item, format: 'TIME | TITLE | SPEAKER'. Speaker may be blank if it's a break or general session.",  "p"),
    ("Speakers — one line per speaker, format: 'NAME | ROLE'. Leave blank if it's a community/peer event with no formal speakers.",  "p"),
    ("Past Action — only for Status = Past. Pick what the link should say: 'View Highlights' (default), 'Watch Recording' (if a video exists), or 'Read Recap'.",  "p"),
    ("Highlights — only for Status = Past. Bullet points of key takeaways, one per line.",  "p"),
    ("Image Filenames — pipe-separated list of photo filenames you will share separately. The web team will upload them to /events/ and they'll appear in the gallery in the order listed.",  "p"),
    ("Notes for Web Team — anything else (logos, sponsor links, special handling).",  "p"),
    ("",                                            "blank"),
    ("Returning the file",                         "h2"),
    ("• Save and email this completed file back to the CiPD web team.",  "p"),
    ("• Send event photos as a separate zip, with filenames matching the 'Image Filenames' column.",  "p"),
]

style_title = Font(name="Calibri", size=18, bold=True, color="00BFA5")
style_h2    = Font(name="Calibri", size=13, bold=True, color="222222")
style_p     = Font(name="Calibri", size=11, color="333333")

for row_idx, (text, kind) in enumerate(instructions, start=1):
    cell = ws2.cell(row=row_idx, column=1, value=text)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if kind == "title":
        cell.font = style_title
        ws2.row_dimensions[row_idx].height = 28
    elif kind == "h2":
        cell.font = style_h2
        ws2.row_dimensions[row_idx].height = 22
    elif kind == "p":
        cell.font = style_p
        ws2.row_dimensions[row_idx].height = 32

# Make Instructions the first sheet so it opens by default
wb.move_sheet("Instructions", offset=-1)

wb.save(OUT_PATH)
print(f"Wrote {OUT_PATH}")
